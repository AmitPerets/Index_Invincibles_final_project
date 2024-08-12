import random
import requests
from bs4 import BeautifulSoup
import pandas as pd
from collections import defaultdict, Counter
import re
import nltk
from nltk.corpus import stopwords
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urljoin
import math
import networkx as nx
import time

nltk.download('stopwords')
stop_words = set(stopwords.words('english'))

def identify_hubs_authorities(article_urls):
    G = nx.DiGraph()
    for url in article_urls:
        links = extract_links_from_page(url, print_links=True)
        for link in links:
            G.add_edge(url, link)
    
    # Print the graph nodes and edges
    print("Nodes:", G.nodes())
    print("Edges:", G.edges())
    
    hubs, authorities = nx.hits(G)
    return hubs, authorities

def apply_hits_algorithm(G):
    hubs, authorities = nx.hits(G)
    return hubs, authorities

def clean_text(text):
    text = re.sub(r'\d+', '', text)  # Remove numbers
    text = re.sub(r'[^\w\s]', '', text)  # Remove punctuation
    words = text.lower().split()
    words = [word for word in words if word not in stop_words]  # Remove stop words
    return words

def get_article_urls(base_url, max_pages=36):
    article_urls = []
    for page in range(1, max_pages + 1):
        response = requests.get(base_url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            links = soup.find_all('a')
            for link in links:
                href = link.get('href')
                if href:
                    full_url = urljoin(base_url, href)
                    article_urls.append(full_url)
    return article_urls

def get_article_details(article_url):
    headers = {'User-Agent': 'Mozilla/5.0'}
    response = requests.get(article_url, headers=headers)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        content = []
        for paragraph in soup.find_all('p'):
            content.append(paragraph.get_text(strip=True))
        content = '\n'.join(content)
        article_title = soup.find('h1').get_text(strip=True) if soup.find('h1') else 'No title'
        return {
            'url': article_url,
            'title': article_title,
            'content': content
        }
    return None

def has_all_keywords(content, keywords):
    for keyword in keywords:
        if keyword.lower() not in content.lower():
            return False
    return True

def keyword_score(content, keywords):
    score = 0
    for keyword in keywords:
        if keyword.lower() in content.lower():
            score += 1
    return score

def find_articles_by_keywords(base_url, keywords, query, max_articles=20, max_pages=36):
    article_urls = get_article_urls(base_url, max_pages)
    matching_articles = []
    shuffled = random.sample(article_urls, k=len(article_urls))
    for url in shuffled:
        if len(matching_articles) >= max_articles:
            break
        details = get_article_details(url)
        if details is not None:
            score = keyword_score(details['content'], keywords)
            if score > 0:  # Adjust this threshold as needed
                matching_articles.append(details)
    return matching_articles

def save_articles_to_excel(articles, filename='articles.xlsx'):
    df = pd.DataFrame(articles, columns=['title', 'url', 'content'])
    df.to_excel(filename, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(filename)
    ws = wb.active

    # Apply styles
    header_font = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(vertical='top', wrap_text=True)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(filename)

def build_inverted_index(articles):
    word_count = Counter()
    for article in articles:
        words = clean_text(article['title']) + clean_text(article['content'])
        word_count.update(words)
    common_words = [word for word, _ in word_count.most_common(15)]

    inverted_index = defaultdict(lambda: [0] * len(articles))
    for idx, article in enumerate(articles):
        words = clean_text(article['title']) + clean_text(article['content'])
        for word in words:
            if word in common_words:
                inverted_index[word][idx] += 1
    return inverted_index

def compute_tf(word, doc):
    return doc.count(word) / len(doc)

def compute_idf(word, docs):
    num_docs_containing_word = sum(1 for doc in docs if word in doc)
    return math.log(len(docs) / (1 + num_docs_containing_word))

def compute_tf_idf(word, doc, docs):
    tf = compute_tf(word, doc)
    idf = compute_idf(word, docs)
    return tf, idf, tf * idf

def calculate_precision_recall(user_relevant):
    tp = sum(user_relevant)  # True Positives
    fp = len(user_relevant) - tp  # False Positives
    fn = 10 - tp  # False Negatives (assuming we have 10 total relevant documents)

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0

    return precision, recall

def extract_links_from_page(url, print_links=False):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        links = [urljoin(url, link.get('href')) for link in soup.find_all('a') if link.get('href')]
        if print_links:
            print(f"Links extracted from {url}:", links)
        return links
    return []

def calculate_hubs_authorities_for_results(article_urls):
    hubs, authorities = identify_hubs_authorities(article_urls)
    
    # Print hubs and authorities scores for each page
    print("\nHubs and Authorities Scores:")
    for url in article_urls:
        hub_score = hubs.get(url, 0)
        authority_score = authorities.get(url, 0)
        print(f"URL: {url}\nHub Score: {hub_score}\nAuthority Score: {authority_score}\n")

def save_results_to_excel(results, filename='results.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Save Top 10 Results
    ws.append(["Top 10 Results"])
    for i, article in enumerate(results['top_10_results']):
        ws.append([f"{i+1}. Title: {article['title']}, URL: {article['url']}"])

    # Save Term Frequencies
    ws.append([])
    ws.append(["Term Frequencies"])
    term_frequencies = results['term_frequencies']
    for r in dataframe_to_rows(term_frequencies, index=True, header=True):
        ws.append(r)

    # Save Number of Terms
    ws.append([])
    ws.append(["Number of Terms"])
    num_terms = results['num_terms']
    for num, doc in num_terms:
        ws.append([num, f"Doc{doc}"])

    # Save Term Frequency Ratios
    ws.append([])
    ws.append(["Term Frequency Ratios"])
    term_frequency_ratios = results['term_frequency_ratios']
    for r in dataframe_to_rows(term_frequency_ratios, index=True, header=True):
        ws.append(r)

    # Save IDF Scores
    ws.append([])
    ws.append(["IDF Scores"])
    idf_scores = results['idf_scores']
    for r in dataframe_to_rows(idf_scores.to_frame(), index=True, header=True):
        ws.append(r)

    # Save TF-IDF Scores
    ws.append([])
    ws.append(["TF-IDF Scores"])
    tf_idf_df = results['tf_idf_df']
    for r in dataframe_to_rows(tf_idf_df, index=True, header=True):
        ws.append(r)

    # Save Precision and Recall
    ws.append([])
    ws.append(["Precision and Recall"])
    ws.append(["User 1 Precision", results['precision1']])
    ws.append(["User 1 Recall", results['recall1']])
    ws.append(["User 2 Precision", results['precision2']])
    ws.append(["User 2 Recall", results['recall2']])
    ws.append(["Average Precision", results['precision']])
    ws.append(["Average Recall", results['recall']])

    # Save Hubs and Authorities Scores
    ws.append([])
    ws.append(["Hubs and Authorities Scores"])
    hubs_authorities = results['hubs_authorities']
    ws.append(["URL", "Hub Score", "Authority Score"])
    for url, scores in hubs_authorities.items():
        ws.append([url, scores['hub_score'], scores['authority_score']])

    wb.save(filename)

def main():
    base_url = 'https://www.ea.com/news'
    queries = [
        {
            'query': 'news',
            'keywords': ['news', 'event', 'update', 'details']
        },
        {
            'query': 'updates',
            'keywords': ['update', 'game', 'performance', 'patch']
        },
        {
            'query': 'players',
            'keywords': ['player', 'statistics', 'ranking', 'top']
        }
    ]

    all_matching_articles = []

    for q in queries:
        if len(all_matching_articles) >= 20:
            break
        max_articles = min(10, 20 - len(all_matching_articles))
        matching_articles = find_articles_by_keywords(base_url, q['keywords'], q['query'], max_articles=max_articles, max_pages=5)
        all_matching_articles.extend(matching_articles)

    if all_matching_articles:
        save_articles_to_excel(all_matching_articles)
        print(f'Found {len(all_matching_articles)} articles. Check out articles.xlsx')

        # Build inverted index for the top 15 most common words
        inverted_index = build_inverted_index(all_matching_articles)

        # Print the inverted index
        for word, vector in inverted_index.items():
            print(f'{word}: {vector}')

        # Create frequency tables
        query_words = ['news', 'event', 'update', 'details']
        docs = [clean_text(details['content']) for details in all_matching_articles]
        
        # Show top 10 results and precision/recall
        top_10_results = all_matching_articles[:10]
        print("\nTop 10 Results:")
        for i, article in enumerate(top_10_results):
            print(f"{i+1}. Title: {article['title']}, URL: {article['url']}")

        term_frequencies = pd.DataFrame({
            word: [clean_text(article['title']).count(word) + clean_text(article['content']).count(word) for article in all_matching_articles]
            for word in query_words
        }, index=[f"doc{i+1}" for i in range(len(docs))])
        print("\nTerm Frequencies:\n", term_frequencies)

        # Calculate number of terms in each document
        num_terms = [(len(clean_text(article['title'])) + len(clean_text(article['content'])), i+1) for i, article in enumerate(all_matching_articles)]
        num_terms_df = pd.DataFrame(num_terms, columns=["Number of Terms", "Doc"])
        print("\nNumber of Terms:\n", num_terms_df)

        doc_lengths = pd.Series([len(clean_text(article['title'])) + len(clean_text(article['content'])) for article in all_matching_articles], index=[f"doc{i+1}" for i in range(len(docs))])
        term_frequency_ratios = pd.DataFrame({
            word: [f"{count}/{doc_lengths[i]}" for i, count in enumerate(term_frequencies[word])]
            for word in query_words
        }, index=[f"doc{i+1}" for i in range(len(docs))])
        print("\nTerm Frequency Ratios:\n", term_frequency_ratios)

        idf_scores = pd.Series({word: compute_idf(word, docs) for word in query_words})
        print("\nIDF Scores:\n", idf_scores)

        # Calculate TF-IDF
        tf_idf_scores = {}
        for word in query_words:
            tf_idf_scores[word] = [compute_tf_idf(word, doc, docs) for doc in docs]

        tf_idf_df = pd.DataFrame({
            word: [f"{tf:.4f}*{idf:.4f}" for tf, idf, _ in tf_idf_scores[word]]
            for word in query_words
        })
        tf_idf_df.index = [f"doc{i+1}" for i in range(len(docs))]

        # Print TF-IDF Scores
        print("\nTF-IDF Scores:\n", tf_idf_df)

        # Assume user relevance feedback
        user1_relevant = [1, 1, 0, 1, 1, 0, 1, 0, 1, 1]  # 1 for relevant, 0 for not relevant
        user2_relevant = [1, 0, 1, 1, 1, 0, 1, 0, 0, 1]  # 1 for relevant, 0 for not relevant

        # Calculate precision and recall
        precision1, recall1 = calculate_precision_recall(user1_relevant)
        precision2, recall2 = calculate_precision_recall(user2_relevant)

        precision = (precision1 + precision2) / 2
        recall = (recall1 + recall2) / 2

        # Display precision and recall
        print("\nPrecision and Recall:")
        print(f"User 1 Precision: {precision1:.4f}")
        print(f"User 1 Recall: {recall1:.4f}")
        print(f"User 2 Precision: {precision2:.4f}")
        print(f"User 2 Recall: {recall2:.4f}")
        print(f"Average Precision: {precision:.4f}")
        print(f"Average Recall: {recall:.4f}")

        # Calculate Hubs and Authorities Scores
        article_urls = [article['url'] for article in all_matching_articles]
        hubs, authorities = identify_hubs_authorities(article_urls)
        hubs_authorities = {url: {'hub_score': hubs.get(url, 0), 'authority_score': authorities.get(url, 0)} for url in article_urls}

        # Print hubs and authorities scores for each page
        print("\nHubs and Authorities Scores:")
        for url in article_urls:
            hub_score = hubs.get(url, 0)
            authority_score = authorities.get(url, 0)
            print(f"URL: {url}\nHub Score: {hub_score}\nAuthority Score: {authority_score}\n")

        # Save results to excel
        results = {
            'top_10_results': top_10_results,
            'tf_idf_df': tf_idf_df,
            'term_frequencies': term_frequencies,
            'term_frequency_ratios': term_frequency_ratios,
            'idf_scores': idf_scores,
            'precision1': precision1,
            'recall1': recall1,
            'precision2': precision2,
            'recall2': recall2,
            'precision': precision,
            'recall': recall,
            'num_terms': num_terms,
            'hubs_authorities': hubs_authorities
        }
        save_results_to_excel(results)

        start_time = time.time()
        matching_articles = find_articles_by_keywords(base_url, q['keywords'], q['query'], max_articles=max_articles, max_pages=5)
        end_time = time.time()
        print(f"Time taken: {end_time - start_time} seconds")
    else:
        print('No articles found matching the given keywords.')

main()
